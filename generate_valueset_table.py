#!/usr/bin/env python3
"""
Script to generate a table of ValueSets from FHIR JSON files.
"""

import json
import glob
import os
from typing import List, Dict, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")

def get_valueset_info(filepath: str) -> Dict[str, Any]:
    """Extract relevant information from a ValueSet JSON file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Get basic info
    name = data.get('name', 'N/A')
    title = data.get('title', name)
    url = data.get('url', 'N/A')
    description = data.get('description', 'N/A')

    # Get concept count from expansion or compose
    concept_count = 0

    # First try expansion (if available)
    if 'expansion' in data and 'contains' in data['expansion']:
        concept_count = len(data['expansion']['contains'])
    # Otherwise count concepts in compose.include
    elif 'compose' in data and 'include' in data['compose']:
        for include in data['compose']['include']:
            if 'concept' in include:
                concept_count += len(include['concept'])
            elif 'valueSet' in include:
                # If it includes other valuesets, we can't easily count
                pass

    # Determine if intensional or extensional
    # Extensional: explicitly lists all concepts (compose.include[].concept exists)
    # Intensional: uses filters or includes entire code systems (compose.include[].filter exists or no concept list)
    definition_type = "Unknown"
    if 'compose' in data and 'include' in data['compose']:
        has_explicit_concepts = False
        has_filters = False
        has_valueset_includes = False

        for include in data['compose']['include']:
            if 'concept' in include and len(include['concept']) > 0:
                has_explicit_concepts = True
            if 'filter' in include:
                has_filters = True
            if 'valueSet' in include:
                has_valueset_includes = True

        if has_filters or has_valueset_includes:
            definition_type = "Intensional"
        elif has_explicit_concepts:
            definition_type = "Extensional"
        else:
            # No concepts and no filters - might be importing entire code system
            definition_type = "Intensional"

    # Get code systems from compose.include
    code_systems = set()
    if 'compose' in data and 'include' in data['compose']:
        for include in data['compose']['include']:
            if 'system' in include:
                code_systems.add(include['system'])

    # Also check expansion for code systems
    if 'expansion' in data and 'contains' in data['expansion']:
        for item in data['expansion']['contains']:
            if 'system' in item:
                code_systems.add(item['system'])

    return {
        'name': name,
        'title': title,
        'url': url,
        'description': description,
        'concept_count': concept_count,
        'definition_type': definition_type,
        'code_systems': sorted(list(code_systems))
    }

def generate_markdown_table(valuesets: List[Dict[str, Any]]) -> str:
    """Generate a markdown table from valueset information."""

    # Sort by name
    valuesets_sorted = sorted(valuesets, key=lambda x: x['name'])

    # Build markdown table
    lines = []
    lines.append("# ValueSet Summary Table\n")
    lines.append("| ValueSet Name | Description | Number of Concepts | Definition Type | Code Systems |")
    lines.append("|---------------|-------------|-------------------|-----------------|--------------|")

    for vs in valuesets_sorted:
        name_link = f"[{vs['title']}]({vs['url']})"
        description = vs['description'].replace('\n', ' ').replace('|', '\\|')
        concept_count = str(vs['concept_count'])
        definition_type = vs['definition_type']
        code_systems = '<br>'.join(vs['code_systems']) if vs['code_systems'] else 'N/A'

        lines.append(f"| {name_link} | {description} | {concept_count} | {definition_type} | {code_systems} |")

    return '\n'.join(lines)

def generate_html_table(valuesets: List[Dict[str, Any]]) -> str:
    """Generate an HTML table from valueset information."""

    # Sort by name
    valuesets_sorted = sorted(valuesets, key=lambda x: x['name'])

    # Build HTML table
    lines = []
    lines.append("<!DOCTYPE html>")
    lines.append("<html>")
    lines.append("<head>")
    lines.append("  <meta charset='utf-8'>")
    lines.append("  <title>ValueSet Summary Table</title>")
    lines.append("  <style>")
    lines.append("    body { font-family: Arial, sans-serif; margin: 20px; }")
    lines.append("    h1 { color: #333; }")
    lines.append("    table { border-collapse: collapse; width: 100%; }")
    lines.append("    th, td { border: 1px solid #ddd; padding: 12px; text-align: left; }")
    lines.append("    th { background-color: #4CAF50; color: white; }")
    lines.append("    tr:nth-child(even) { background-color: #f2f2f2; }")
    lines.append("    tr:hover { background-color: #ddd; }")
    lines.append("    a { color: #0066cc; text-decoration: none; }")
    lines.append("    a:hover { text-decoration: underline; }")
    lines.append("  </style>")
    lines.append("</head>")
    lines.append("<body>")
    lines.append("  <h1>ValueSet Summary Table</h1>")
    lines.append("  <table>")
    lines.append("    <tr>")
    lines.append("      <th>ValueSet Name</th>")
    lines.append("      <th>Description</th>")
    lines.append("      <th>Number of Concepts</th>")
    lines.append("      <th>Definition Type</th>")
    lines.append("      <th>Code Systems</th>")
    lines.append("    </tr>")

    for vs in valuesets_sorted:
        name_link = f"<a href='{vs['url']}'>{vs['title']}</a>"
        description = vs['description'].replace('<', '&lt;').replace('>', '&gt;')
        concept_count = str(vs['concept_count'])
        definition_type = vs['definition_type']
        code_systems = '<br>'.join(vs['code_systems']) if vs['code_systems'] else 'N/A'

        lines.append("    <tr>")
        lines.append(f"      <td>{name_link}</td>")
        lines.append(f"      <td>{description}</td>")
        lines.append(f"      <td>{concept_count}</td>")
        lines.append(f"      <td>{definition_type}</td>")
        lines.append(f"      <td>{code_systems}</td>")
        lines.append("    </tr>")

    lines.append("  </table>")
    lines.append("</body>")
    lines.append("</html>")

    return '\n'.join(lines)

def generate_excel_table(valuesets: List[Dict[str, Any]], filename: str = 'valueset_table.xlsx'):
    """Generate an Excel spreadsheet from valueset information."""

    if not OPENPYXL_AVAILABLE:
        print("Skipping Excel generation - openpyxl not installed")
        return

    # Sort by name
    valuesets_sorted = sorted(valuesets, key=lambda x: x['name'])

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "ValueSet Summary"

    # Define header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Write headers
    headers = ["ValueSet Name", "URL", "Description", "Number of Concepts", "Definition Type", "Code Systems"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data
    for row_idx, vs in enumerate(valuesets_sorted, 2):
        # Column A: ValueSet Name (Title)
        ws.cell(row=row_idx, column=1, value=vs['title'])

        # Column B: URL (as hyperlink)
        url_cell = ws.cell(row=row_idx, column=2, value=vs['url'])
        url_cell.hyperlink = vs['url']
        url_cell.font = Font(color="0066CC", underline="single")

        # Column C: Description
        desc_cell = ws.cell(row=row_idx, column=3, value=vs['description'])
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Column D: Number of Concepts
        ws.cell(row=row_idx, column=4, value=vs['concept_count'])

        # Column E: Definition Type
        ws.cell(row=row_idx, column=5, value=vs['definition_type'])

        # Column F: Code Systems (with line breaks)
        code_systems_text = '\n'.join(vs['code_systems']) if vs['code_systems'] else 'N/A'
        cs_cell = ws.cell(row=row_idx, column=6, value=code_systems_text)
        cs_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 50

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save workbook
    wb.save(filename)
    print(f"Excel table saved to: {filename}")

def main():
    """Main function to process ValueSet files and generate tables."""

    # Find all ValueSet JSON files
    pattern = "output/ValueSet-*.json"
    files = glob.glob(pattern)

    if not files:
        print(f"No files found matching pattern: {pattern}")
        return

    print(f"Found {len(files)} ValueSet files")

    # Process each file
    valuesets = []
    for filepath in files:
        try:
            vs_info = get_valueset_info(filepath)
            valuesets.append(vs_info)
            print(f"Processed: {vs_info['name']}")
        except Exception as e:
            print(f"Error processing {filepath}: {e}")

    # Generate markdown table
    markdown_table = generate_markdown_table(valuesets)
    with open('valueset_table.md', 'w', encoding='utf-8') as f:
        f.write(markdown_table)
    print("\nMarkdown table saved to: valueset_table.md")

    # Generate HTML table
    html_table = generate_html_table(valuesets)
    with open('valueset_table.html', 'w', encoding='utf-8') as f:
        f.write(html_table)
    print("HTML table saved to: valueset_table.html")

    # Generate Excel table
    generate_excel_table(valuesets, 'valueset_table.xlsx')

    print(f"\nTotal ValueSets processed: {len(valuesets)}")

if __name__ == "__main__":
    main()
